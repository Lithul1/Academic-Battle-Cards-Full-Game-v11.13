# -*- coding: utf-8 -*-
"""Write the first editions into the catalog: standard tier to '1st Editions',
ultra tier to a new 'Special Editions' tab. Run after sync.py."""
import json, sys
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

XL = sys.argv[1] if len(sys.argv)>1 else 'Academic_Battle_Cards_Catalog_SYNCED.xlsx'
F  = json.load(open('firsteds.json'))
ARCH = {r['id']: r['arch'] for r in json.load(open('fe_arch.json'))}

COLS = ['id','name','deck','tier','hp','archetype','accent','base',
        'passive_name','passive_text',
        'atk_name','atk_text','atk_cost','atk_dmg','atk_label','atk_extra_json',
        'blk_name','blk_text','blk_cost','blk_block','blk_label','blk_extra_json']
SKIP_A = {'n','t','cost','dmg','label'}
SKIP_B = {'n','t','cost','block','label'}

def row_for(c):
    a,b = c['atk'], c['blk']
    ea = {k:v for k,v in a.items() if k not in SKIP_A}
    eb = {k:v for k,v in b.items() if k not in SKIP_B}
    return [c['id'], c['name'], c.get('deck',''), c.get('tier','standard'), c.get('hp'),
            ARCH.get(c['id'], c.get('archetype','')), c.get('accent',''),
            ', '.join(c.get('base') or []),
            (c.get('passive') or {}).get('name',''), (c.get('passive') or {}).get('text',''),
            a.get('n'), a.get('t'), a.get('cost'), a.get('dmg'), a.get('label'),
            json.dumps(ea, ensure_ascii=False) if ea else '',
            b.get('n'), b.get('t'), b.get('cost'), b.get('block'), b.get('label'),
            json.dumps(eb, ensure_ascii=False) if eb else '']

F_HD  = PatternFill("solid", fgColor="1F3864")
F_SP  = PatternFill("solid", fgColor="6B4E8A")
ALT   = PatternFill("solid", fgColor="F2F2F2")
thin  = Side(style="thin", color="BFBFBF"); BOX = Border(left=thin,right=thin,top=thin,bottom=thin)
WIDTH = [18,24,13,10,6,15,9,20, 22,52, 24,54,7,7,9,30, 24,54,7,7,9,30]

def write_sheet(wb, title, rows, headfill):
    if title in wb.sheetnames: del wb[title]
    ws = wb.create_sheet(title)
    ws.append(COLS)
    for i in range(1, len(COLS)+1):
        x = ws.cell(1,i); x.fill=headfill
        x.font=Font(name='Arial',bold=True,color="FFFFFF",size=10)
        x.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    ws.row_dimensions[1].height=34; ws.freeze_panes="C2"
    for r,c in enumerate(rows, 2):
        for i,v in enumerate(row_for(c), 1): ws.cell(r,i,value=v)
        for i in range(1,len(COLS)+1):
            x=ws.cell(r,i); x.font=Font(name='Arial',size=10); x.border=BOX
            x.alignment=Alignment(vertical="top", wrap_text=(i in (10,12,18,16,22)))
            if r%2==0: x.fill=ALT
        ws.row_dimensions[r].height=46
    for i,w in enumerate(WIDTH,1): ws.column_dimensions[get_column_letter(i)].width=w
    ws.auto_filter.ref=f"A1:{get_column_letter(len(COLS))}{len(rows)+1}"
    return ws

wb = load_workbook(XL)
std  = [c for c in F if c.get('tier','standard') != 'ultra']
ultra= [c for c in F if c.get('tier','standard') == 'ultra']
write_sheet(wb, '1st Editions', std, F_HD)
ws2 = write_sheet(wb, 'Special Editions', ultra, F_SP)
ws2.cell(len(ultra)+3, 1, value='Ultra-rare first editions live here. Add a row per card; '
        'the sync writes tier:"ultra" cards to this tab and everything else to 1st Editions.'
       ).font = Font(name='Arial', size=10, italic=True, color="6B4E8A")
wb.save(XL)
print(f"1st Editions: {len(std)} rows | Special Editions: {len(ultra)} rows")
print("sheets:", wb.sheetnames)
